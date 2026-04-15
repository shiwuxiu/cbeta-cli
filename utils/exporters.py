"""Multi-format exporters for CBETA CLI.

支持导出格式：
- JSON (默认)
- CSV
- Excel (xlsx)
- Markdown
- HTML
"""

import json
import csv
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from datetime import datetime


def export_to_json(data: Union[Dict, List], file_path: Path) -> None:
    """导出为 JSON 格式."""
    file_path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8"
    )


def export_to_csv(data: Union[Dict, List], file_path: Path) -> None:
    """导出为 CSV 格式."""
    # 确保数据是列表形式
    if isinstance(data, dict):
        if "results" in data:
            rows = data["results"]
        else:
            rows = [data]
    else:
        rows = data

    if not rows:
        file_path.write_text("", encoding="utf-8")
        return

    # 确定列名（从第一个记录提取）
    if isinstance(rows[0], dict):
        columns = list(rows[0].keys())
    else:
        columns = ["value"]
        rows = [{"value": r} for r in rows]

    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            # 处理嵌套值
            flat_row = {}
            for k, v in row.items():
                if isinstance(v, (dict, list)):
                    flat_row[k] = json.dumps(v, ensure_ascii=False)
                else:
                    flat_row[k] = v
            writer.writerow(flat_row)


def export_to_excel(data: Union[Dict, List], file_path: Path) -> None:
    """导出为 Excel 格式."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise RuntimeError("需要安装 openpyxl: pip install openpyxl")

    # 确保数据是列表形式
    if isinstance(data, dict):
        if "results" in data:
            rows = data["results"]
        else:
            rows = [data]
    else:
        rows = data

    wb = Workbook()
    ws = wb.active
    ws.title = "CBETA Data"

    if not rows:
        wb.save(file_path)
        return

    # 写入表头
    if isinstance(rows[0], dict):
        columns = list(rows[0].keys())
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 写入数据
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name)
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(file_path)


def export_to_markdown(data: Union[Dict, List], file_path: Path, title: str = "CBETA Data") -> None:
    """导出为 Markdown 格式."""
    lines = [f"# {title}", "", f"导出时间: {datetime.now().isoformat()}", ""]

    # 确保数据是列表形式
    if isinstance(data, dict):
        if "results" in data:
            rows = data["results"]
            # 添加摘要信息
            if "num_found" in data:
                lines.append(f"**找到 {data['num_found']} 条结果**")
                lines.append("")
        else:
            # 单条记录
            lines.append("## 数据详情")
            lines.append("")
            for k, v in data.items():
                if isinstance(v, (dict, list)):
                    lines.append(f"- **{k}**: `{json.dumps(v, ensure_ascii=False)}`")
                else:
                    lines.append(f"- **{k}**: {v}")
            file_path.write_text("\n".join(lines), encoding="utf-8")
            return
    else:
        rows = data

    if not rows:
        lines.append("*无数据*")
        file_path.write_text("\n".join(lines), encoding="utf-8")
        return

    # 生成表格
    if isinstance(rows[0], dict):
        columns = list(rows[0].keys())

        # 表头
        lines.append("| " + " | ".join(columns) + " |")
        lines.append("| " + " | ".join(["---"] * len(columns)) + " |")

        # 数据行
        for row in rows:
            values = []
            for col in columns:
                v = row.get(col)
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)[:50] + "..." if len(json.dumps(v)) > 50 else json.dumps(v, ensure_ascii=False)
                elif v is None:
                    v = ""
                values.append(str(v).replace("|", "\\|"))
            lines.append("| " + " | ".join(values) + " |")

    file_path.write_text("\n".join(lines), encoding="utf-8")


def export_to_html(data: Union[Dict, List], file_path: Path, title: str = "CBETA Data") -> None:
    """导出为 HTML 格式."""
    html_parts = [
        "<!DOCTYPE html>",
        "<html lang='zh-CN'>",
        "<head>",
        "<meta charset='utf-8'>",
        f"<title>{title}</title>",
        "<style>",
        "body { font-family: 'Microsoft YaHei', sans-serif; margin: 20px; }",
        "table { border-collapse: collapse; width: 100%; }",
        "th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
        "th { background-color: #4CAF50; color: white; }",
        "tr:nth-child(even) { background-color: #f2f2f2; }",
        ".metadata { background: #f0f0f0; padding: 10px; margin-bottom: 20px; }",
        "</style>",
        "</head>",
        "<body>",
        f"<h1>{title}</h1>",
        f"<div class='metadata'>导出时间: {datetime.now().isoformat()}</div>",
    ]

    # 确保数据是列表形式
    if isinstance(data, dict):
        if "results" in data:
            rows = data["results"]
            if "num_found" in data:
                html_parts.append(f"<p><strong>找到 {data['num_found']} 条结果</strong></p>")
        else:
            # 单条记录
            html_parts.append("<h2>数据详情</h2>")
            html_parts.append("<table>")
            for k, v in data.items():
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                html_parts.append(f"<tr><th>{k}</th><td>{v}</td></tr>")
            html_parts.append("</table>")
            html_parts.extend(["</body>", "</html>"])
            file_path.write_text("\n".join(html_parts), encoding="utf-8")
            return
    else:
        rows = data

    if not rows:
        html_parts.append("<p><em>无数据</em></p>")
        html_parts.extend(["</body>", "</html>"])
        file_path.write_text("\n".join(html_parts), encoding="utf-8")
        return

    # 生成表格
    if isinstance(rows[0], dict):
        columns = list(rows[0].keys())

        html_parts.append("<table>")
        html_parts.append("<tr>" + "".join(f"<th>{c}</th>" for c in columns) + "</tr>")

        for row in rows:
            html_parts.append("<tr>")
            for col in columns:
                v = row.get(col)
                if isinstance(v, (dict, list)):
                    v = f"<code>{json.dumps(v, ensure_ascii=False)[:100]}</code>"
                elif v is None:
                    v = ""
                html_parts.append(f"<td>{v}</td>")
            html_parts.append("</tr>")

        html_parts.append("</table>")

    html_parts.extend(["</body>", "</html>"])
    file_path.write_text("\n".join(html_parts), encoding="utf-8")


def get_exporter(format: str) -> callable:
    """获取导出函数."""
    exporters = {
        "json": export_to_json,
        "csv": export_to_csv,
        "excel": export_to_excel,
        "xlsx": export_to_excel,
        "markdown": export_to_markdown,
        "md": export_to_markdown,
        "html": export_to_html,
    }

    format_lower = format.lower()
    if format_lower not in exporters:
        raise ValueError(f"不支持格式 '{format}'，支持的格式: {list(exporters.keys())}")

    return exporters[format_lower]